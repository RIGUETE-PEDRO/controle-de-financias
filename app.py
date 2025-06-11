# app.py
import os
import sqlite3
import io
import openpyxl
from flask import Flask, request, jsonify, render_template, make_response
from flask_cors import CORS
from datetime import datetime
from database import * # Importa todas as funções do nosso arquivo database

app = Flask(__name__)
CORS(app) 
iniciar_db()

@app.route('/')
def index():
    return render_template('index.html')

# --- ROTAS DE TRANSAÇÃO ---
@app.route('/api/transacoes', methods=['GET'])
def api_obter_transacoes():
    usuario = request.args.get('usuario')
    if not usuario: return jsonify({"erro": "Usuário é obrigatório"}), 400
    transacoes = obter_transacoes_por_usuario(usuario)
    return jsonify(transacoes)

@app.route('/api/transacoes', methods=['POST'])
def api_adicionar_transacao():
    dados = request.get_json()
    if not all(k in dados for k in ['descricao', 'valor', 'tipo', 'usuario', 'data']):
        return jsonify({"erro": "Dados incompletos"}), 400
    adicionar_transacao(dados['descricao'], float(dados['valor']), dados['tipo'], dados['data'], dados['usuario'], dados.get('meta_id'))
    return jsonify({"sucesso": "Transação adicionada!"}), 201

@app.route('/api/transacoes/<int:transacao_id>', methods=['DELETE'])
def api_remover_transacao(transacao_id):
    remover_transacao(transacao_id)
    return jsonify({"sucesso": "Transação removida!"}), 200

# --- ROTAS DE META ---
@app.route('/api/metas', methods=['GET'])
def api_obter_metas():
    usuario = request.args.get('usuario')
    if not usuario: return jsonify({"erro": "Usuário é obrigatório"}), 400
    metas = obter_metas_por_usuario(usuario)
    return jsonify(metas)

@app.route('/api/metas', methods=['POST'])
def api_criar_meta():
    dados = request.get_json()
    if not all(k in dados for k in ['nome', 'percentual', 'usuario']):
        return jsonify({"erro": "Dados incompletos"}), 400
    criar_meta(dados['nome'], float(dados['percentual']), dados['usuario'])
    return jsonify({"sucesso": "Meta criada!"}), 201

@app.route('/api/metas/<int:meta_id>', methods=['DELETE'])
def api_remover_meta(meta_id):
    remover_meta(meta_id)
    return jsonify({"sucesso": "Meta removida!"}), 200
# app.py
# Adicione estas importações no início do arquivo, junto com as outras
from openpyxl.styles import Font, PatternFill

# ... (resto do arquivo) ...

# --- ROTA DE RELATÓRIO ATUALIZADA COM CORES ---
@app.route('/api/relatorio', methods=['GET'])
def api_gerar_relatorio():
    usuario = request.args.get('usuario')
    data_de = request.args.get('de')
    data_ate = request.args.get('ate')

    if not all([usuario, data_de, data_ate]):
        return jsonify({"erro": "Usuário e período são obrigatórios"}), 400

    transacoes = obter_transacoes_por_periodo(usuario, data_de, data_ate)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f"Relatório de {usuario}"
    
    # --- Estilos de Cor ---
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F4F4F", end_color="4F4F4F", fill_type="solid")
    green_font = Font(color="008000", bold=True)
    red_font = Font(color="FF0000", bold=True)

    # Cabeçalho
    headers = ["Data", "Tipo", "Descrição", "Valor"]
    sheet.append(headers)
    for cell in sheet[1]: # Pega a primeira linha
        cell.font = header_font
        cell.fill = header_fill

    total_entradas = 0
    total_saidas = 0

    # Preenche as linhas
    for t in transacoes:
        valor = t['valor']
        data_formatada = datetime.fromisoformat(t['data']).strftime('%d/%m/%Y')
        tipo = t['tipo'].capitalize()
        
        row_data = [data_formatada, tipo, t['descricao'], valor]
        sheet.append(row_data)
        
        # Pega a última linha adicionada para aplicar o estilo
        last_row = sheet.max_row
        valor_cell = sheet[f'D{last_row}']
        
        if t['tipo'] == 'entrada':
            total_entradas += valor
            valor_cell.font = green_font
        else:
            total_saidas += valor
            valor_cell.font = red_font

    # Adiciona resumo no final
    sheet.append([])
    sheet.append(["", "", "Total Entradas:", total_entradas])
    sheet.append(["", "", "Total Saídas:", total_saidas])
    sheet.append(["", "", "Saldo do Período:", total_entradas - total_saidas])

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    response = make_response(buffer.getvalue())
    response.headers['Content-Disposition'] = f'attachment; filename=relatorio_{usuario}_{data_de}_a_{data_ate}.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response

# ... (resto do arquivo)
# if __name__ == '__main__': ...

# app.py
# ... (outras importações) ...
# Importe a nova função
from database import obter_gastos_por_categoria

# ... (outras rotas) ...

# --- NOVA ROTA PARA DADOS DO GRÁFICO DE PIZZA ---
@app.route('/api/relatorio/pizza', methods=['GET'])
def api_gerar_relatorio_pizza():
    usuario = request.args.get('usuario')
    mes = request.args.get('mes') # Espera um formato 'YYYY-MM'

    if not all([usuario, mes]):
        return jsonify({"erro": "Usuário e mês são obrigatórios"}), 400
    
    dados_grafico = obter_gastos_por_categoria(usuario, mes)
    return jsonify(dados_grafico)

# ... (resto do arquivo) ...